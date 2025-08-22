import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
from openpyxl.utils import get_column_letter

# ---------------------------
# Utility Functions
# ---------------------------

def create_uid(row):
    return f"{str(row['TOWNSHIP']).strip()}{str(row['REGISTRATION NUMBER']).strip()}{str(row['PATIENT NAME']).strip()}"

def format_worksheet_dates(worksheet, df):
    date_format = 'yyyy-mm-dd'
    for idx, col in enumerate(df.columns):
        max_length = len(str(col)) + 2
        column_letter = get_column_letter(idx + 1)
        for row_idx in range(len(df)):
            cell = worksheet.cell(row=row_idx + 2, column=idx + 1)
            cell_value = df.iloc[row_idx, idx]
            if isinstance(cell_value, (datetime, pd.Timestamp)):
                cell.value = cell_value.date()
                cell.number_format = date_format
                cell_length = 10
            else:
                cell_length = len(str(cell_value))
            max_length = max(max_length, cell_length + 2)
        worksheet.column_dimensions[column_letter].width = max_length

# ---------------------------
# Streamlit UI
# ---------------------------
st.title("📊 TB Excel Processing App")
st.write("Upload your Excel file and get the processed outputs.")

uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

# Logging container
log_container = st.empty()

# Initialize log session state
if "log_text" not in st.session_state:
    st.session_state.log_text = ""

def log(message):
    st.session_state.log_text += f"{message}\n"
    log_container.text_area("Processing Logs", value=st.session_state.log_text, height=300)

# ---------------------------
# Main Processing
# ---------------------------
if uploaded_file:
    log(f"✅ File uploaded: {uploaded_file.name}")

    # Define sheet types
    column_style_sheets = ['Kutkai','Kunlong','Muse','Namhkan','Namhsan','Namtu','Hseni','Lashio','Laukkaing']
    row_style_sheets = ['Kyauktaw','Myebon','Minbya','Mrauk-U','Buthidaung','Rathedaung','Maungdaw','Paletwa']

    try:
        workbook = load_workbook(uploaded_file, read_only=True)
        all_sheet_names = workbook.sheetnames
        workbook.close()

        existing_column_sheets = [s for s in column_style_sheets if s in all_sheet_names]
        existing_row_sheets = [s for s in row_style_sheets if s in all_sheet_names]

        log(f"📄 Sheets detected: Column-style {len(existing_column_sheets)}, Row-style {len(existing_row_sheets)}")

        # Prepare in-memory files
        row_output_file = BytesIO()
        column_output_file = BytesIO()
        final_output_file = BytesIO()

        # ---------------------------
        # Process Row-Style Sheets
        # ---------------------------
        row_summary_data = []
        row_sheets_data = {}

        for sheet_name in existing_row_sheets:
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
            df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
            df['TOWNSHIP'] = sheet_name
            row_sheets_data[sheet_name] = df.copy()
            row_summary_data.append(df)
            log(f"   Processed row-style sheet: {sheet_name}")

        if row_summary_data:
            combined_row_summary = pd.concat(row_summary_data, ignore_index=True)
            with pd.ExcelWriter(row_output_file, engine='openpyxl') as writer:
                combined_row_summary.to_excel(writer, sheet_name='Summary', index=False)
                for sheet_name, df in row_sheets_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            row_output_file.seek(0)
            st.download_button("Download Row-Style Sheets", data=row_output_file,
                               file_name="Row_Style_Sheets.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            log("✅ Row-style sheets processed")

        # ---------------------------
        # Process Column-Style Sheets
        # ---------------------------
        summary_data_list = []
        column_sheets_data = {}

        for sheet_name in existing_column_sheets:
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
            df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
            column_sheets_data[sheet_name] = df.copy()

            # Simplified: extract registration & visit date
            cols_needed = [c for c in df.columns if 'registration' in c.lower() or 'visit' in c.lower()]
            summary_df = df[cols_needed].copy()
            summary_df['SOURCE_SHEET'] = sheet_name
            summary_df['TOWNSHIP'] = sheet_name
            summary_data_list.append(summary_df)
            log(f"   Processed column-style sheet: {sheet_name}")

        if summary_data_list:
            combined_summary = pd.concat(summary_data_list, ignore_index=True)
            with pd.ExcelWriter(column_output_file, engine='openpyxl') as writer:
                combined_summary.to_excel(writer, sheet_name='Summary', index=False)
                for sheet_name, df in column_sheets_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            column_output_file.seek(0)
            st.download_button("Download Column-Style Sheets & Summary", data=column_output_file,
                               file_name="Summary_and_Column_Sheets.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            log("✅ Column-style sheets processed")

        # ---------------------------
        # Final Restructured File
        # ---------------------------
        with pd.ExcelWriter(final_output_file, engine='openpyxl') as writer:
            if summary_data_list:
                combined_summary.to_excel(writer, sheet_name='Summary', index=False)
            for sheet_name, df in column_sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            for sheet_name, df in row_sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        final_output_file.seek(0)
        st.download_button("Download Final Restructured File", data=final_output_file,
                           file_name="Final_Restructured_Data.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        log("🎉 All files processed successfully!")

    except Exception as e:
        log(f"❌ Error: {e}")
